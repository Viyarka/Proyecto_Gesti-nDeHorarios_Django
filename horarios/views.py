import io
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import (
    ConfiguracionFranjaForm,
    DisponibilidadProfesorForm,
    GenerarGlobalForm,
    GenerarHorarioForm,
    HorarioFilterForm,
    HorarioForm,
    PerfilAlumnoForm,
    SesionForm,
)
from .models import (
    Asignatura,
    AuditLog,
    DisponibilidadProfesor,
    FranjaHoraria,
    Horario,
    Notificacion,
    PerfilAlumno,
    Profesor,
    Sesion,
)
from .motor import (
    detectar_conflictos_alumno,
    detectar_conflictos_asignaturas,
    generar_horario_basico,
    generar_horarios_globales,
    validar_horario,
)


def pertenece_a_grupo(user, nombre_grupo):
    return user.is_superuser or user.groups.filter(name=nombre_grupo).exists()


def es_direccion(user):
    return pertenece_a_grupo(user, "Decanato") or pertenece_a_grupo(user, "IT")


def registrar_auditoria(request, accion, modelo, objeto_id, descripcion, valor_anterior="", valor_nuevo=""):
    AuditLog.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        accion=accion,
        modelo=modelo,
        objeto_id=str(objeto_id),
        descripcion=descripcion,
        valor_anterior=valor_anterior,
        valor_nuevo=valor_nuevo,
        ip=request.META.get("REMOTE_ADDR"),
    )


def queryset_sesiones_filtrado(params):
    qs = Sesion.objects.select_related(
        "horario", "horario__titulacion", "horario__curso", "asignatura", "profesor", "aula", "grupo", "franja"
    )
    form = HorarioFilterForm(params or None)
    if form.is_valid():
        data = form.cleaned_data
        if data.get("anio_academico"):
            qs = qs.filter(horario__anio_academico__icontains=data["anio_academico"])
        if data.get("semestre"):
            qs = qs.filter(horario__semestre=data["semestre"])
        if data.get("titulacion"):
            qs = qs.filter(horario__titulacion__nombre__icontains=data["titulacion"])
        if data.get("curso"):
            qs = qs.filter(horario__curso__numero=data["curso"])
        if data.get("grupo"):
            qs = qs.filter(grupo__nombre__icontains=data["grupo"])
        if data.get("profesor"):
            qs = qs.filter(Q(profesor__nombre__icontains=data["profesor"]) | Q(profesor__codigo__icontains=data["profesor"]))
        if data.get("asignatura"):
            qs = qs.filter(Q(asignatura__nombre__icontains=data["asignatura"]) | Q(asignatura__codigo__icontains=data["asignatura"]))
    return form, qs.order_by("horario__anio_academico", "horario__semestre", "horario__titulacion__codigo", "horario__curso__numero", "franja__dia", "franja__hora_inicio")


@login_required
def dashboard(request):
    horarios = Horario.objects.all()
    sesiones = Sesion.objects.all()
    visitados = request.session.get("horarios_visitados", [])
    horarios_visitados = list(Horario.objects.select_related("titulacion", "curso").filter(pk__in=visitados))
    posicion = {pk: indice for indice, pk in enumerate(visitados)}
    horarios_visitados.sort(key=lambda h: posicion.get(h.pk, 999))

    contexto = {
        "total_horarios": horarios.count(),
        "total_sesiones": sesiones.count(),
        "horarios_revision": horarios.filter(estado="REVISION").count(),
        "horarios_aprobados": horarios.filter(estado="APROBADO").count(),
        "notificaciones": Notificacion.objects.filter(usuario=request.user, leida=False)[:5],
        "ultimos_horarios": horarios_visitados[:5],
        "total_profesores": Profesor.objects.count(),
        "total_asignaturas": Asignatura.objects.count(),
    }
    return render(request, "horarios/dashboard.html", contexto)


class HorarioListView(LoginRequiredMixin, ListView):
    model = Horario
    template_name = "horarios/horario_list.html"
    context_object_name = "horarios"
    paginate_by = 50

    def get_queryset(self):
        qs = Horario.objects.select_related("titulacion", "curso").annotate(num_sesiones=Count("sesiones", distinct=True))
        self.form = HorarioFilterForm(self.request.GET or None)
        if self.form.is_valid():
            data = self.form.cleaned_data
            if data.get("anio_academico"):
                qs = qs.filter(anio_academico__icontains=data["anio_academico"])
            if data.get("semestre"):
                qs = qs.filter(semestre=data["semestre"])
            if data.get("titulacion"):
                qs = qs.filter(titulacion__nombre__icontains=data["titulacion"])
            if data.get("curso"):
                qs = qs.filter(curso__numero=data["curso"])
            if data.get("grupo"):
                qs = qs.filter(sesiones__grupo__nombre__icontains=data["grupo"]).distinct()
            if data.get("profesor"):
                qs = qs.filter(Q(sesiones__profesor__nombre__icontains=data["profesor"]) | Q(sesiones__profesor__codigo__icontains=data["profesor"])).distinct()
            if data.get("asignatura"):
                qs = qs.filter(Q(sesiones__asignatura__nombre__icontains=data["asignatura"]) | Q(sesiones__asignatura__codigo__icontains=data["asignatura"])).distinct()
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["form"] = self.form
        return ctx


@login_required
def horario_detalle(request, pk):
    horario = get_object_or_404(Horario.objects.select_related("titulacion", "curso"), pk=pk)
    sesiones = horario.sesiones.select_related("asignatura", "profesor", "aula", "grupo", "franja")
    resultado = validar_horario(horario)

    visitados = request.session.get("horarios_visitados", [])
    if horario.pk in visitados:
        visitados.remove(horario.pk)
    request.session["horarios_visitados"] = [horario.pk] + visitados[:4]
    request.session.modified = True

    dias = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]
    franjas_unicas = (
        horario.sesiones
        .values("franja__hora_inicio", "franja__hora_fin")
        .distinct()
        .order_by("franja__hora_inicio", "franja__hora_fin")
    )
    tabla = []
    for franja in franjas_unicas:
        inicio = franja["franja__hora_inicio"]
        fin = franja["franja__hora_fin"]
        fila = {"inicio": inicio, "fin": fin, "dias": []}
        for dia in dias:
            celda = sesiones.filter(
                franja__dia=dia,
                franja__hora_inicio=inicio,
                franja__hora_fin=fin,
            ).first()
            fila["dias"].append(celda)
        tabla.append(fila)

    return render(request, "horarios/horario_detail.html", {"horario": horario, "tabla": tabla, "dias": dias, "resultado": resultado})


class HorarioCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Horario
    form_class = HorarioForm
    template_name = "horarios/form.html"
    permission_required = "horarios.add_horario"

    def form_valid(self, form):
        response = super().form_valid(form)
        registrar_auditoria(self.request, "CREATE", "Horario", self.object.pk, "Creación de horario.")
        messages.success(self.request, "Horario creado correctamente.")
        return response


class HorarioUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Horario
    form_class = HorarioForm
    template_name = "horarios/form.html"
    permission_required = "horarios.change_horario"

    def form_valid(self, form):
        anterior = Horario.objects.get(pk=self.object.pk).estado
        response = super().form_valid(form)
        registrar_auditoria(self.request, "UPDATE", "Horario", self.object.pk, "Edición de horario.", anterior, self.object.estado)
        messages.success(self.request, "Horario actualizado.")
        return response


class HorarioDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Horario
    template_name = "horarios/confirm_delete.html"
    success_url = reverse_lazy("horario_lista")
    permission_required = "horarios.delete_horario"


class SesionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Sesion
    form_class = SesionForm
    template_name = "horarios/form.html"
    permission_required = "horarios.add_sesion"

    def form_valid(self, form):
        response = super().form_valid(form)
        registrar_auditoria(self.request, "CREATE", "Sesion", self.object.pk, "Creación manual de sesión.")
        messages.success(self.request, "Sesión creada correctamente.")
        return response

    def get_success_url(self):
        return self.object.horario.get_absolute_url()


class SesionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Sesion
    form_class = SesionForm
    template_name = "horarios/form.html"
    permission_required = "horarios.change_sesion"

    def form_valid(self, form):
        response = super().form_valid(form)
        registrar_auditoria(self.request, "UPDATE", "Sesion", self.object.pk, "Edición manual de sesión.")
        messages.success(self.request, "Sesión actualizada y validada por el motor de reglas.")
        return response

    def get_success_url(self):
        return self.object.horario.get_absolute_url()


class SesionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Sesion
    template_name = "horarios/confirm_delete.html"
    permission_required = "horarios.delete_sesion"

    def get_success_url(self):
        return self.object.horario.get_absolute_url()


@login_required
@user_passes_test(es_direccion)
def generar_horario(request):
    if request.method == "POST":
        form = GenerarHorarioForm(request.POST)
        if form.is_valid():
            horario = form.cleaned_data["horario"]
            try:
                creadas = generar_horario_basico(horario, form.cleaned_data["sobrescribir"])
                registrar_auditoria(request, "GENERATION", "Horario", horario.pk, f"Generación automática individual: {creadas} sesiones.")
                messages.success(request, f"Generación finalizada: {creadas} sesiones creadas.")
                return redirect(horario)
            except ValidationError as exc:
                messages.error(request, exc.message if hasattr(exc, "message") else exc)
    else:
        form = GenerarHorarioForm()
    return render(request, "horarios/form.html", {"form": form, "titulo": "Generar un horario"})


@login_required
@user_passes_test(es_direccion)
def generar_global(request):
    if request.method == "POST":
        form = GenerarGlobalForm(request.POST)
        if form.is_valid():
            resultado = generar_horarios_globales(
                anio_academico=form.cleaned_data["anio_academico"],
                semestre=form.cleaned_data["semestre"],
                sobrescribir=form.cleaned_data["sobrescribir"],
            )
            registrar_auditoria(request, "GENERATION", "Horario", "GLOBAL", f"Generación global: {resultado.creadas} creadas, {resultado.omitidas} omitidas.")
            if resultado.errores:
                messages.warning(request, f"Generación con avisos: {resultado.creadas} sesiones creadas y {resultado.omitidas} no colocadas.")
            else:
                messages.success(request, f"Generación global finalizada: {resultado.creadas} sesiones creadas.")
            return render(request, "horarios/generacion_resultado.html", {"resultado": resultado, "form": form})
    else:
        form = GenerarGlobalForm()
    return render(request, "horarios/generar_global.html", {"form": form})


@login_required
def validar_horario_view(request, pk):
    horario = get_object_or_404(Horario, pk=pk)
    resultado = validar_horario(horario, exigir_horas=True)
    registrar_auditoria(request, "VALIDATION", "Horario", horario.pk, "Validación completa de horario.")
    return render(request, "horarios/validacion.html", {"horario": horario, "resultado": resultado})


@login_required
@user_passes_test(es_direccion)
def cambiar_estado(request, pk, nuevo_estado):
    horario = get_object_or_404(Horario, pk=pk)
    anterior = horario.estado
    if nuevo_estado == "APROBADO":
        resultado = validar_horario(horario, exigir_horas=True)
        if not resultado["es_valido"]:
            messages.error(request, "No se puede aprobar un horario con conflictos o con horas lectivas incompletas.")
            return redirect("validar_horario", pk=horario.pk)
        horario.aprobado_por = request.user
        horario.fecha_aprobacion = timezone.now()

    horario.estado = nuevo_estado
    horario.save(update_fields=["estado", "aprobado_por", "fecha_aprobacion", "actualizado"])
    registrar_auditoria(request, "WORKFLOW", "Horario", horario.pk, "Cambio de estado.", anterior, nuevo_estado)
    messages.success(request, f"Estado cambiado a {horario.get_estado_display()}.")
    return redirect(horario)


@login_required
def carga_profesor(request):
    profesor = Profesor.objects.filter(user=request.user).first()
    if not profesor and not request.user.is_superuser:
        raise PermissionDenied("Tu usuario no está vinculado a un profesor.")
    sesiones = Sesion.objects.select_related("horario", "asignatura", "grupo", "aula", "franja", "profesor")
    if profesor:
        sesiones = sesiones.filter(profesor=profesor)
    return render(request, "horarios/carga_profesor.html", {"sesiones": sesiones, "profesor": profesor})


@login_required
def disponibilidad_profesor(request):
    profesor = Profesor.objects.filter(user=request.user).first()
    if request.method == "POST":
        form = DisponibilidadProfesorForm(request.POST)
        if form.is_valid():
            disp = form.save()
            registrar_auditoria(request, "UPDATE", "DisponibilidadProfesor", disp.pk, "Cambio de disponibilidad.")
            messages.success(request, "Disponibilidad guardada e incorporada al historial.")
            return redirect("disponibilidad_profesor")
    else:
        inicial = {"profesor": profesor} if profesor else {}
        form = DisponibilidadProfesorForm(initial=inicial)

    disponibilidades = DisponibilidadProfesor.objects.select_related("profesor", "franja")
    if profesor:
        disponibilidades = disponibilidades.filter(profesor=profesor)

    return render(request, "horarios/disponibilidad.html", {"form": form, "disponibilidades": disponibilidades})


@login_required
def horario_alumno(request):
    perfil = PerfilAlumno.objects.filter(user=request.user).first()
    if not perfil:
        messages.warning(request, "No tienes perfil de alumno creado. Usa el usuario de demo alumno / Alumno12345.")
        return redirect("dashboard")

    sesiones = Sesion.objects.filter(Q(grupo=perfil.grupo) | Q(asignatura__in=perfil.asignaturas.all())).select_related(
        "asignatura", "profesor", "aula", "grupo", "franja", "horario"
    ).distinct()

    conflictos = detectar_conflictos_alumno(perfil)
    return render(request, "horarios/horario_alumno.html", {"perfil": perfil, "sesiones": sesiones, "conflictos": conflictos})


@login_required
def matricula_alumno(request):
    perfil, _ = PerfilAlumno.objects.get_or_create(user=request.user)
    conflictos = []
    if request.method == "POST":
        form = PerfilAlumnoForm(request.POST, instance=perfil)
        if form.is_valid():
            asignaturas = form.cleaned_data["asignaturas"]
            conflictos = detectar_conflictos_asignaturas(asignaturas)
            if conflictos:
                messages.error(request, "La selección tiene solapamientos. Revisa las asignaturas indicadas antes de guardar.")
            else:
                form.save()
                registrar_auditoria(request, "UPDATE", "PerfilAlumno", perfil.pk, "Actualización de matrícula sin conflictos.")
                messages.success(request, "Matrícula actualizada sin solapamientos.")
                return redirect("horario_alumno")
    else:
        form = PerfilAlumnoForm(instance=perfil)
    return render(request, "horarios/matricula.html", {"form": form, "conflictos": conflictos})


@login_required
def notificaciones(request):
    qs = Notificacion.objects.filter(usuario=request.user)
    if request.method == "POST":
        qs.update(leida=True)
        messages.success(request, "Notificaciones marcadas como leídas.")
        return redirect("notificaciones")
    return render(request, "horarios/notificaciones.html", {"notificaciones": qs})


@login_required
@user_passes_test(es_direccion)
def auditoria(request):
    logs = AuditLog.objects.select_related("usuario")[:100]
    return render(request, "horarios/auditoria.html", {"logs": logs})


@login_required
@user_passes_test(es_direccion)
def configuracion_franjas(request):
    if request.method == "POST":
        form = ConfiguracionFranjaForm(request.POST)
        if form.is_valid():
            franja = form.save()
            registrar_auditoria(request, "UPDATE", "FranjaHoraria", franja.pk, "Configuración de franja.")
            messages.success(request, "Franja guardada.")
            return redirect("configuracion_franjas")
    else:
        form = ConfiguracionFranjaForm()
    franjas = FranjaHoraria.objects.all()
    return render(request, "horarios/configuracion_franjas.html", {"form": form, "franjas": franjas})


def escribir_sesiones_excel(sesiones, titulo="Horario académico"):
    sesiones = list(sesiones)
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    columnas = ["Año", "Semestre", "Titulación", "Curso", "Día", "Inicio", "Fin", "Grupo", "Asignatura", "Código", "Profesor", "Aula"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    ws.cell(row=1, column=1, value=titulo)
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="002855")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
    ws.cell(row=2, column=1, value=f"Total de sesiones: {len(sesiones)}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="334155")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    ws.append(columnas)
    header_row = 3
    header_fill = PatternFill("solid", fgColor="0B3A66")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for s in sesiones:
        ws.append([
            s.horario.anio_academico,
            s.horario.get_semestre_display(),
            s.horario.titulacion.nombre,
            f"{s.horario.curso.numero}º",
            s.franja.get_dia_display(),
            s.franja.hora_inicio.strftime("%H:%M"),
            s.franja.hora_fin.strftime("%H:%M"),
            s.grupo.nombre,
            s.asignatura.nombre,
            s.asignatura.codigo,
            s.profesor.codigo,
            s.aula.nombre,
        ])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    widths = [13, 18, 30, 8, 13, 10, 10, 12, 34, 16, 16, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(columnas))}{ws.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def sesiones_horario_matriz(horario):
    """Prepara una matriz semanal solo con las sesiones del horario actual."""
    dias = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]
    sesiones = horario.sesiones.select_related("asignatura", "profesor", "aula", "grupo", "franja")
    franjas_unicas = (
        sesiones.values("franja__hora_inicio", "franja__hora_fin")
        .distinct()
        .order_by("franja__hora_inicio", "franja__hora_fin")
    )

    filas = []
    for franja in franjas_unicas:
        inicio = franja["franja__hora_inicio"]
        fin = franja["franja__hora_fin"]
        fila = {"inicio": inicio, "fin": fin, "dias": []}
        for dia in dias:
            sesion = sesiones.filter(
                franja__dia=dia,
                franja__hora_inicio=inicio,
                franja__hora_fin=fin,
            ).first()
            fila["dias"].append(sesion)
        filas.append(fila)
    return dias, filas


def escribir_horario_excel(horario):
    """Exporta el horario que se está viendo en formato tabla semanal."""
    dias, filas = sesiones_horario_matriz(horario)
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    columnas = ["Franja"] + [dia.title() for dia in dias]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    ws.cell(row=1, column=1, value=f"Horario {horario}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=18, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="002855")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    total = horario.sesiones.count()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
    ws.cell(row=2, column=1, value=f"{horario.get_semestre_display()} · {horario.titulacion.nombre} · {horario.curso.numero}º · Total de sesiones: {total}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="334155")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    ws.append(columnas)
    header_row = 3
    header_fill = PatternFill("solid", fgColor="0B3A66")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for fila in filas:
        valores = [f"{fila['inicio']:%H:%M} - {fila['fin']:%H:%M}"]
        for sesion in fila["dias"]:
            if sesion:
                valores.append(
                    f"{sesion.asignatura.nombre}\n"
                    f"{sesion.grupo.nombre} · {sesion.profesor.codigo}\n"
                    f"{sesion.aula.nombre}"
                )
            else:
                valores.append("Libre")
        ws.append(valores)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].font = Font(bold=True, color="002855")
        row[0].fill = PatternFill("solid", fgColor="EAF1F8")
        if row[0].row % 2 == 0:
            for cell in row[1:]:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    ws.column_dimensions["A"].width = 16
    for idx in range(2, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 32
    for row_idx in range(4, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "B4"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@login_required
def exportar_excel(request, pk):
    horario = get_object_or_404(
        Horario.objects.select_related("titulacion", "curso"),
        pk=pk,
    )
    output = escribir_horario_excel(horario)
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="horario_{horario.pk}.xlsx"'
    return response


def escribir_sesiones_pdf(titulo, sesiones):
    sesiones = list(sesiones)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=22, rightMargin=22, topMargin=22, bottomMargin=22)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(titulo, styles["Title"]),
        Paragraph(f"Informe generado desde el sistema de gestión de horarios. Total de sesiones: {len(sesiones)}", styles["Normal"]),
        Spacer(1, 14),
    ]
    data = [["Año", "Semestre", "Titulación", "Curso", "Día", "Hora", "Grupo", "Asignatura", "Profesor", "Aula"]]
    for s in sesiones:
        data.append([
            s.horario.anio_academico,
            s.horario.get_semestre_display(),
            s.horario.titulacion.codigo,
            f"{s.horario.curso.numero}º",
            s.franja.get_dia_display(),
            f"{s.franja.hora_inicio:%H:%M}-{s.franja.hora_fin:%H:%M}",
            s.grupo.nombre,
            s.asignatura.nombre[:36],
            s.profesor.codigo,
            s.aula.nombre,
        ])
    table = Table(data, repeatRows=1, colWidths=[58, 78, 58, 38, 70, 70, 55, 170, 75, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002855")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return output


def escribir_horario_pdf(horario):
    """Exporta únicamente el horario actual en formato tabla semanal."""
    dias, filas = sesiones_horario_matriz(horario)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=22, rightMargin=22, topMargin=22, bottomMargin=22)
    styles = getSampleStyleSheet()
    titulo = f"Horario {horario}"
    subtitulo = f"{horario.get_semestre_display()} · {horario.titulacion.nombre} · {horario.curso.numero}º · Total de sesiones: {horario.sesiones.count()}"
    story = [Paragraph(titulo, styles["Title"]), Paragraph(subtitulo, styles["Normal"]), Spacer(1, 12)]

    data = [["Franja"] + [dia.title() for dia in dias]]
    for fila in filas:
        row = [f"{fila['inicio']:%H:%M}<br/>{fila['fin']:%H:%M}"]
        for sesion in fila["dias"]:
            if sesion:
                contenido = (
                    f"<b>{sesion.asignatura.nombre}</b><br/>"
                    f"{sesion.grupo.nombre} · {sesion.profesor.codigo}<br/>"
                    f"{sesion.aula.nombre}"
                )
            else:
                contenido = "<font color='#64748B'>Libre</font>"
            row.append(Paragraph(contenido, styles["BodyText"]))
        data.append(row)

    table = Table(data, repeatRows=1, colWidths=[62, 142, 142, 142, 142, 142])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002855")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#EAF1F8")),
        ("TEXTCOLOR", (0, 1), (0, -1), colors.HexColor("#002855")),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return output


@login_required
def exportar_pdf(request, pk):
    horario = get_object_or_404(
        Horario.objects.select_related("titulacion", "curso"),
        pk=pk,
    )
    output = escribir_horario_pdf(horario)
    return FileResponse(output, as_attachment=True, filename=f"horario_{horario.pk}.pdf")


@login_required
def informes(request):
    form, sesiones = queryset_sesiones_filtrado(request.GET)
    registrar_auditoria(request, "REPORT", "Sesion", "FILTRO", "Consulta de informe dinámico.")
    querystring = urlencode(request.GET)
    return render(request, "horarios/informes.html", {"form": form, "sesiones": sesiones[:500], "querystring": querystring})


@login_required
def exportar_informe_excel(request):
    _, sesiones = queryset_sesiones_filtrado(request.GET)
    output = escribir_sesiones_excel(sesiones, titulo="Informe dinámico de horarios")
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="informe_horarios.xlsx"'
    return response


@login_required
def exportar_informe_pdf(request):
    _, sesiones = queryset_sesiones_filtrado(request.GET)
    output = escribir_sesiones_pdf("Informe dinámico de horarios", sesiones[:500])
    return FileResponse(output, as_attachment=True, filename="informe_horarios.pdf")


@login_required
def api_sesiones(request):
    _, sesiones = queryset_sesiones_filtrado(request.GET)
    data = [
        {
            "id": s.id,
            "anio_academico": s.horario.anio_academico,
            "semestre": s.horario.semestre,
            "titulacion": s.horario.titulacion.nombre,
            "curso": s.horario.curso.numero,
            "grupo": s.grupo.nombre,
            "dia": s.franja.dia,
            "hora_inicio": s.franja.hora_inicio.strftime("%H:%M"),
            "hora_fin": s.franja.hora_fin.strftime("%H:%M"),
            "asignatura": s.asignatura.nombre,
            "codigo_asignatura": s.asignatura.codigo,
            "transversal": s.asignatura.es_transversal,
            "profesor": s.profesor.codigo,
            "aula": s.aula.nombre,
        }
        for s in sesiones[:1000]
    ]
    return JsonResponse({"results": data})
